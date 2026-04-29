import streamlit as st
import pandas as pd
import re
from io import BytesIO
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

st.set_page_config(page_title="Гүйлгээний алдаа шалгах", layout="wide", page_icon="🔍")

st.title("🔍 Гүйлгээний алдаа шалгах")
st.caption("TRR XML Report файл upload хийж алдаатай гүйлгээг шалгана")

# ── helpers ──────────────────────────────────────────────────────────────────
def parse_num(v):
    v = str(v).replace(" ₮", "").replace("₮", "").replace(",", "").strip()
    try:
        return float(v)
    except:
        return 0.0

def mls_base(v):
    m = re.match(r"^(\d+)", str(v))
    return m.group(1) if m else str(v)

def load_file(uploaded):
    raw = uploaded.read()
    # HTML-based XLS (XML/HTML table)
    try:
        soup = BeautifulSoup(raw.decode("utf-8-sig"), "html.parser")
        table = soup.find("table")
        if table:
            rows = table.find_all("tr")
            headers = [th.get_text(strip=True) for th in rows[0].find_all(["th", "td"])]
            data = []
            for row in rows[1:]:
                cells = [td.get_text(strip=True) for td in row.find_all("td")]
                if cells:
                    data.append(cells)
            return pd.DataFrame(data, columns=headers)
    except:
        pass
    # Real XLS/XLSX
    try:
        return pd.read_excel(BytesIO(raw), dtype=str)
    except:
        return pd.read_excel(BytesIO(raw), engine="openpyxl", dtype=str)

def check_errors(df):
    # ── 1. ₮ replace ────────────────────────────────────────────────────────
    for col in df.columns:
        df[col] = df[col].astype(str).str.replace(" ₮", "", regex=False).str.replace("₮", "", regex=False)

    # ── 2. Тооцоолол ────────────────────────────────────────────────────────
    df["_comm"] = df["Total Commission"].apply(parse_num)
    df["_sold"]  = df["Зарагдсан үнэ"].apply(parse_num)
    df["Шимтгэлийн хувь"] = df.apply(
        lambda r: round(r["_comm"] / r["_sold"] * 100, 2) if r["_sold"] > 0 else 0.0, axis=1
    )

    # ── 3. MLS ID ────────────────────────────────────────────────────────────
    df["MLS_ID"] = df["Бүртгэлийн дугаар"].apply(mls_base)

    # ── 4. Алдааны баганууд ──────────────────────────────────────────────────
    errors_list = []   # алдааны мөрүүд хадгалах

    # 4-a. TRR ID давхардал
    dup_trr = df.duplicated("TRR ID", keep=False)
    df["Алдаа_TRR"] = dup_trr.map({True: "Давхардсан", False: ""})

    # 4-b. AgentID өөр дээрээ хаасан: 1 MLS ID дээр нэг AgentID 2+ удаа
    agent_cnt = df.groupby(["MLS_ID", "AgentID"]).size().reset_index(name="_n")
    agent_dup = set(
        zip(agent_cnt[agent_cnt["_n"] >= 2]["MLS_ID"],
            agent_cnt[agent_cnt["_n"] >= 2]["AgentID"])
    )
    df["Алдаа_Агент"] = df.apply(
        lambda r: "Агент өөр дээрээ хаасан"
        if (r["MLS_ID"], r["AgentID"]) in agent_dup else "", axis=1
    )

    # 4-c. Шимтгэл зөрсөн
    VALID = {
        # (Шилжүүлэгийн төрөл, TRR Type) : set of valid pct
        ("Түрээс",   "Listing and Selling TRR"): {20.0, 50.0, 90.0},
        ("Түрээс",   "Listing TRR"):             {10.0, 25.0, 45.0},
        ("Худалдах", "Listing and Selling TRR"): {3.0, 5.0},
        ("Худалдах", "Listing TRR"):             {1.5, 2.5},
    }

    def shimtgel_error(row):
        key = (row["Шилжүүлэгийн төрөл"], row["TRR Type"])
        if key not in VALID:
            return ""
        pct = round(float(row["Шимтгэлийн хувь"]), 1)
        valid_set = VALID[key]
        if pct not in valid_set:
            return "Шимтгэл зөрсөн"
        return ""

    df["Алдаа_Шимтгэл"] = df.apply(shimtgel_error, axis=1)

    # ── 5. Нийт алдаа багана ────────────────────────────────────────────────
    def combine(row):
        parts = [row["Алдаа_TRR"], row["Алдаа_Агент"], row["Алдаа_Шимтгэл"]]
        parts = [p for p in parts if p]
        return " | ".join(parts)

    df["Алдаа"] = df.apply(combine, axis=1)

    # ── 6. Туслах баганууд арилгах ───────────────────────────────────────────
    df.drop(columns=["_comm", "_sold", "MLS_ID"], inplace=True)

    return df

def to_excel(df_all, df_err):
    wb = Workbook()

    thin = Side(style="thin", color="D0D0D0")
    bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)

    def make_sheet(wb, df, title, tab_color):
        ws = wb.active if title == "Бүх мэдээлэл" else wb.create_sheet(title)
        ws.title = title
        ws.tab_color = tab_color
        ws.freeze_panes = "A2"
        ws.sheet_view.showGridLines = False

        # Header
        hdr_fill = PatternFill("solid", fgColor="1E3A5F")
        err_fill = PatternFill("solid", fgColor="7B1010")
        for ci, col in enumerate(df.columns, 1):
            c = ws.cell(row=1, column=ci, value=col)
            c.font      = Font(name="Arial", bold=True, size=10, color="FFFFFF")
            c.fill      = err_fill if col == "Алдаа" else hdr_fill
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border    = bdr
        ws.row_dimensions[1].height = 22
        ws.auto_filter.ref = ws.dimensions

        # Rows
        err_row_fill  = PatternFill("solid", fgColor="FFF0F0")
        even_fill     = PatternFill("solid", fgColor="F5F8FF")
        odd_fill      = PatternFill("solid", fgColor="FFFFFF")
        red_font      = Font(name="Arial", size=9, color="CC0000", bold=True)
        normal_font   = Font(name="Arial", size=9)

        for ri, (_, row) in enumerate(df.iterrows(), 2):
            has_err = bool(row.get("Алдаа", ""))
            bg = err_row_fill if has_err else (even_fill if ri % 2 == 0 else odd_fill)
            for ci, val in enumerate(row, 1):
                col_name = df.columns[ci - 1]
                c = ws.cell(row=ri, column=ci, value=val)
                c.font   = red_font if (col_name == "Алдаа" and has_err) else normal_font
                c.fill   = bg
                c.alignment = Alignment(vertical="center", wrap_text=False)
                c.border = bdr
                if col_name == "Шимтгэлийн хувь" and val:
                    try:
                        c.value       = float(val)
                        c.number_format = "0.00%"
                        c.value       = float(val) / 100
                    except:
                        pass
            ws.row_dimensions[ri].height = 15

        # Колонны өргөн
        col_w = {
            "TRR ID": 12, "TRR Type": 22, "Үл хөдлөх хөрөнгийн хаяг": 30,
            "Шилжүүлэгийн төрөл": 16, "Orig. List Date": 14, "Анхны жагсаалтын үнэ": 16,
            "Зарагдсан өдөр": 14, "Зарагдсан үнэ": 16, "Payment Amount": 16,
            "Payment Date": 14, "Payments Received": 16, "Оффисын нэр": 20,
            "AgentID": 14, "Агент": 24, "Бүртгэлийн дугаар": 18,
            "Дүүрэг": 12, "Total Commission": 16, "# of Agents": 10,
            "Total Received": 16, "Total Outstanding": 16, "Last Submission Date": 18,
            "Buyers": 22, "Банк": 20, "Currency": 10, "Market Segment": 14,
            "First Payment": 12, "Шимтгэлийн хувь": 14, "Алдаа": 30,
            "Алдаа_TRR": 16, "Алдаа_Агент": 24, "Алдаа_Шимтгэл": 18,
        }
        for ci, col in enumerate(df.columns, 1):
            ws.column_dimensions[get_column_letter(ci)].width = col_w.get(col, 14)

    # Sheet 1: бүх мэдээлэл
    make_sheet(wb, df_all, "Бүх мэдээлэл", "1E3A5F")
    # Sheet 2: зөвхөн алдаатай
    make_sheet(wb, df_err, "Алдаатай гүйлгээ", "CC0000")

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf

# ── UI ───────────────────────────────────────────────────────────────────────
uploaded = st.file_uploader(
    "📂 XLS / XLSX файл сонгоно уу",
    type=["xls", "xlsx"],
    help="Unegui.mn TRR XML Report файл"
)

if uploaded:
    with st.spinner("Файл уншиж байна..."):
        try:
            df_raw = load_file(uploaded)
        except Exception as e:
            st.error(f"Файл уншихад алдаа гарлаа: {e}")
            st.stop()

    st.success(f"✅ {len(df_raw):,} мөр уншлаа — {len(df_raw.columns)} багана")

    with st.spinner("Алдаа шалгаж байна..."):
        df = check_errors(df_raw.copy())

    df_err = df[df["Алдаа"] != ""].reset_index(drop=True)

    # ── KPI cards ─────────────────────────────────────────────────────────
    n_total   = len(df)
    n_err     = len(df_err)
    n_dup_trr = (df["Алдаа_TRR"] != "").sum()
    n_agent   = (df["Алдаа_Агент"] != "").sum()
    n_shimtg  = (df["Алдаа_Шимтгэл"] != "").sum()

    st.markdown("---")
    c1, c2, c3, c4, c5 = st.columns(5)
    c1.metric("Нийт гүйлгээ",      f"{n_total:,}")
    c2.metric("🔴 Нийт алдаа",      f"{n_err:,}",    delta=f"{n_err/n_total*100:.1f}%", delta_color="inverse")
    c3.metric("🔁 Давхардсан TRR",  f"{n_dup_trr:,}")
    c4.metric("👤 Агент өөрт хаасан", f"{n_agent:,}")
    c5.metric("💰 Шимтгэл зөрсөн",  f"{n_shimtg:,}")
    st.markdown("---")

    # ── Tabs ──────────────────────────────────────────────────────────────
    tab1, tab2 = st.tabs(["🔴 Алдаатай гүйлгээ", "📋 Бүх мэдээлэл"])

    show_cols = [
        "TRR ID", "TRR Type", "Шилжүүлэгийн төрөл",
        "Үл хөдлөх хөрөнгийн хаяг", "Зарагдсан үнэ",
        "Total Commission", "Шимтгэлийн хувь",
        "AgentID", "Агент", "Бүртгэлийн дугаар",
        "Алдаа_TRR", "Алдаа_Агент", "Алдаа_Шимтгэл", "Алдаа"
    ]
    show_cols = [c for c in show_cols if c in df.columns]

    with tab1:
        if n_err == 0:
            st.success("✅ Алдаатай гүйлгээ олдсонгүй!")
        else:
            st.dataframe(
                df_err[show_cols].style.applymap(
                    lambda v: "background-color:#ffe0e0;color:#cc0000;font-weight:bold" if v else "",
                    subset=["Алдаа"]
                ),
                use_container_width=True,
                height=500,
            )

    with tab2:
        st.dataframe(df[show_cols], use_container_width=True, height=500)

    # ── Download ──────────────────────────────────────────────────────────
    st.markdown("---")
    with st.spinner("Excel файл бэлтгэж байна..."):
        excel_buf = to_excel(df, df_err)

    st.download_button(
        label=f"📥 Excel татах — {n_err:,} алдаатай гүйлгээ",
        data=excel_buf,
        file_name="trr_aldaa_shalgah.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        type="primary",
    )
    st.caption("Excel файлд 2 sheet байна: 'Бүх мэдээлэл' ба 'Алдаатай гүйлгээ'")

else:
    st.info("👆 XLS файл upload хийнэ үү")
    with st.expander("ℹ️ Шалгадаг алдаануудын тайлбар"):
        st.markdown("""
| Алдааны төрөл | Тайлбар |
|---|---|
| **Давхардсан** | TRR ID давхардсан байна |
| **Агент өөр дээрээ хаасан** | 1 MLS ID дээр нэг AgentID 2+ удаа бүртгэгдсэн |
| **Шимтгэл зөрсөн** | Шимтгэлийн хувь зөвшөөрөгдсөн утгаас зөрсөн |

**Зөвшөөрөгдсөн шимтгэлийн хувь:**

| Шилжүүлэгийн төрөл | TRR Type | Зөв хувь |
|---|---|---|
| Түрээс | Listing and Selling TRR | 20%, 50%, 90% |
| Түрээс | Listing TRR | 10%, 25%, 45% |
| Худалдах | Listing and Selling TRR | 3%, 5% |
| Худалдах | Listing TRR | 1.5%, 2.5% |
""")
